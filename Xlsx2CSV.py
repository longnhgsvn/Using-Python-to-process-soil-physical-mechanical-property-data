# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Any, Iterable, Optional


# Ensure Vietnamese output works on Windows consoles with legacy codepages.
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


_TCVN3_UNICODE_PAIRS: list[tuple[str, str]] = [
    # Nguồn mapping: vietnamese-conversion (TCVN3 chuẩn VietKey).
    ("Aµ", "À"),
    ("A¸", "Á"),
    ("¢", "Â"),
    ("A·", "Ã"),
    ("EÌ", "È"),
    ("EÐ", "É"),
    ("£", "Ê"),
    ("I×", "Ì"),
    ("IÝ", "Í"),
    ("Oß", "Ò"),
    ("Oã", "Ó"),
    ("¤", "Ô"),
    ("Oâ", "Õ"),
    ("Uï", "Ù"),
    ("Uó", "Ú"),
    ("Yý", "Ý"),
    ("µ", "à"),
    ("¸", "á"),
    ("©", "â"),
    ("·", "ã"),
    ("Ì", "è"),
    ("Ð", "é"),
    ("ª", "ê"),
    ("×", "ì"),
    ("Ý", "í"),
    ("ß", "ò"),
    ("ã", "ó"),
    ("«", "ô"),
    ("â", "õ"),
    ("ï", "ù"),
    ("ó", "ú"),
    ("ý", "ý"),
    ("¡", "Ă"),
    ("§", "Đ"),
    ("IÜ", "Ĩ"),
    ("Uò", "Ũ"),
    ("¥", "Ơ"),
    ("¦", "Ư"),
    ("¨", "ă"),
    ("®", "đ"),
    ("Ü", "ĩ"),
    ("ò", "ũ"),
    ("¬", "ơ"),
    ("­", "ư"),
    ("A¹", "Ạ"),
    ("A¶", "Ả"),
    ("¢Ê", "Ấ"),
    ("¢Ç", "Ầ"),
    ("¢È", "Ẩ"),
    ("¢É", "Ẫ"),
    ("¢Ë", "Ậ"),
    ("¡¾", "Ắ"),
    ("¡»", "Ằ"),
    ("¡¼", "Ẳ"),
    ("¡½", "Ẵ"),
    ("¡Æ", "Ặ"),
    ("EÑ", "Ẹ"),
    ("EÎ", "Ẻ"),
    ("EÏ", "Ẽ"),
    ("£Õ", "Ế"),
    ("£Ò", "Ề"),
    ("£Ó", "Ể"),
    ("£Ô", "Ễ"),
    ("£Ö", "Ệ"),
    ("IØ", "Ỉ"),
    ("IÞ", "Ị"),
    ("Oä", "Ọ"),
    ("Oá", "Ỏ"),
    ("¤è", "Ố"),
    ("¤å", "Ồ"),
    ("¤æ", "Ổ"),
    ("¤ç", "Ỗ"),
    ("¤é", "Ộ"),
    ("¥í", "Ớ"),
    ("¥ê", "Ờ"),
    ("¥ë", "Ở"),
    ("¥ì", "Ỡ"),
    ("¥î", "Ợ"),
    ("Uô", "Ụ"),
    ("Uñ", "Ủ"),
    ("¦ø", "Ứ"),
    ("¦õ", "Ừ"),
    ("¦ö", "Ử"),
    ("¦÷", "Ữ"),
    ("¦ù", "Ự"),
    ("Yú", "Ỳ"),
    ("Yþ", "Ỵ"),
    ("Yû", "Ỷ"),
    ("Yü", "Ỹ"),
    ("¹", "ạ"),
    ("¶", "ả"),
    ("Ê", "ấ"),
    ("Ç", "ầ"),
    ("È", "ẩ"),
    ("É", "ẫ"),
    ("Ë", "ậ"),
    ("¾", "ắ"),
    ("»", "ằ"),
    ("¼", "ẳ"),
    ("½", "ẵ"),
    ("Æ", "ặ"),
    ("Ñ", "ẹ"),
    ("Î", "ẻ"),
    ("Ï", "ẽ"),
    ("Õ", "ế"),
    ("Ò", "ề"),
    ("Ó", "ể"),
    ("Ô", "ễ"),
    ("Ö", "ệ"),
    ("Ø", "ỉ"),
    ("Þ", "ị"),
    ("ä", "ọ"),
    ("á", "ỏ"),
    ("è", "ố"),
    ("å", "ồ"),
    ("æ", "ổ"),
    ("ç", "ỗ"),
    ("é", "ộ"),
    ("í", "ớ"),
    ("ê", "ờ"),
    ("ë", "ở"),
    ("ì", "ỡ"),
    ("î", "ợ"),
    ("ô", "ụ"),
    ("ñ", "ủ"),
    ("ø", "ứ"),
    ("õ", "ừ"),
    ("ö", "ử"),
    ("÷", "ữ"),
    ("ù", "ự"),
    ("ú", "ỳ"),
    ("þ", "ỵ"),
    ("û", "ỷ"),
    ("ü", "ỹ"),
]

_TCVN3_TO_UNICODE = sorted(_TCVN3_UNICODE_PAIRS, key=lambda p: len(p[0]), reverse=True)
_TCVN3_MARKERS_RE = re.compile(r"[µ¸©·¡§ÌÐÝÓ¨®Ü¬­¹¶Ê»¼½Æ]")
_UNICODE_VN_RE = re.compile(
    r"[áàảãạăắằẳẵặâấầẩẫậéèẻẽẹêếềểễệíìỉĩịóòỏõọôốồổỗộơớờởỡợúùủũụưứừửữựýỳỷỹỵđĐ]",
    re.IGNORECASE,
)


def looks_like_tcvn3(text: str) -> bool:
    # TCVN3 thường chứa các ký tự đặc thù như µ¸©...; Unicode "chuẩn" sẽ có áàả...
    if not text:
        return False
    if _TCVN3_MARKERS_RE.search(text) is None:
        return False
    # Nếu vừa có marker TCVN3 vừa có Unicode tiếng Việt, vẫn có thể là chuỗi trộn -> vẫn cho phép chuyển.
    return True


def tcvn3_to_unicode(text: str) -> str:
    if not text:
        return text
    # 2-phase replacement để tránh đè lẫn nhau (ví dụ: '¢' và '¢Ê')
    out = text
    for i, (src, _dst) in enumerate(_TCVN3_TO_UNICODE):
        out = out.replace(src, f"::{i}::")
    for i, (_src, dst) in enumerate(_TCVN3_TO_UNICODE):
        out = out.replace(f"::{i}::", dst)
    return out


def is_tcvn3_font(font_name: Optional[str]) -> bool:
    if not font_name:
        return False
    f = font_name.strip().lower()
    # Common TCVN3 fonts in Excel: ".VnTime", ".VnArial", ".VnCourier", ...
    return ("tcvn3" in f) or f.startswith(".vn")


def safe_filename_piece(name: str) -> str:
    name = name.strip()
    name = re.sub(r"[\\/:*?\"<>|]+", "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:180] if name else "sheet"


def has_any_data(rows: Iterable[Iterable[Any]]) -> bool:
    for r in rows:
        for v in r:
            if v is None:
                continue
            if isinstance(v, str):
                if v.strip() != "":
                    return True
            else:
                return True
    return False


def write_csv(path: Path, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for r in rows:
            w.writerow(["" if v is None else v for v in r])


def convert_xlsx(file_path: Path, out_dir: Path) -> int:
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "Thiếu thư viện openpyxl. Cài bằng: pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(file_path, data_only=True)
    written = 0

    for ws in wb.worksheets:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            continue

        rows: list[list[Any]] = []
        for r in range(1, max_row + 1):
            row: list[Any] = []
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if isinstance(v, str):
                    font_name = getattr(cell.font, "name", None)
                    if is_tcvn3_font(font_name) or looks_like_tcvn3(v):
                        v = tcvn3_to_unicode(v)
                row.append(v)
            rows.append(row)

        if not has_any_data(rows):
            continue

        out_name = f"{file_path.stem}_{safe_filename_piece(ws.title)}.csv"
        write_csv(out_dir / out_name, rows)
        written += 1

    return written


def convert_xls(file_path: Path, out_dir: Path) -> int:
    try:
        import xlrd  # type: ignore
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "Thiếu thư viện xlrd để đọc .xls. Cài bằng: pip install xlrd==1.2.0"
        ) from e

    try:
        book = xlrd.open_workbook(file_path.as_posix(), formatting_info=True)
    except TypeError:
        # xlrd>=2.0.0 removes xls support; user must install 1.2.0
        raise RuntimeError(
            "Phiên bản xlrd hiện tại không hỗ trợ .xls. Hãy cài: pip install xlrd==1.2.0"
        )

    written = 0
    for sheet in book.sheets():
        if sheet.nrows == 0 or sheet.ncols == 0:
            continue

        rows: list[list[Any]] = []
        for r in range(sheet.nrows):
            row: list[Any] = []
            for c in range(sheet.ncols):
                v = sheet.cell_value(r, c)
                if v == "":
                    v = None
                font_name: Optional[str] = None
                try:
                    xf = book.xf_list[sheet.cell_xf_index(r, c)]
                    font = book.font_list[xf.font_index]
                    font_name = getattr(font, "name", None)
                except Exception:
                    font_name = None

                if isinstance(v, str):
                    if is_tcvn3_font(font_name) or looks_like_tcvn3(v):
                        v = tcvn3_to_unicode(v)
                row.append(v)
            rows.append(row)

        if not has_any_data(rows):
            continue

        out_name = f"{file_path.stem}_{safe_filename_piece(sheet.name)}.csv"
        write_csv(out_dir / out_name, rows)
        written += 1

    return written


def main(argv: list[str]) -> int:
    script_dir = Path(__file__).resolve().parent
    root = script_dir.parent

    p = argparse.ArgumentParser(
        description="Chuyển toàn bộ .xls/.xlsx trong dataXLSX sang CSV; đổi TCVN3 -> UTF-8 theo font."
    )
    p.add_argument(
        "--input",
        default=str(root / "dataXLSX"),
        help="Thư mục chứa file .xls/.xlsx (mặc định: <root>/dataXLSX)",
    )
    p.add_argument(
        "--output",
        default=str(root / "dataCSV"),
        help="Thư mục xuất CSV (mặc định: <root>/dataCSV)",
    )
    args = p.parse_args(argv)

    in_dir = Path(args.input).resolve()
    out_dir = Path(args.output).resolve()

    if not in_dir.exists() or not in_dir.is_dir():
        print(f"Không tìm thấy thư mục input: {in_dir}", file=sys.stderr)
        return 2

    files = sorted([*in_dir.glob("*.xlsx"), *in_dir.glob("*.xls")])
    if not files:
        print(f"Không có file .xls/.xlsx trong: {in_dir}")
        return 0

    total_sheets = 0
    failed: list[tuple[Path, str]] = []
    for f in files:
        try:
            if f.suffix.lower() == ".xlsx":
                total_sheets += convert_xlsx(f, out_dir)
            else:
                total_sheets += convert_xls(f, out_dir)
            print(f"OK: {f.name}")
        except Exception as e:  # noqa: BLE001
            failed.append((f, str(e)))
            print(f"FAIL: {f.name} -> {e}", file=sys.stderr)

    print(f"Đã xuất {total_sheets} file CSV vào: {out_dir}")
    if failed:
        print("Các file lỗi:", file=sys.stderr)
        for f, msg in failed:
            print(f"- {f.name}: {msg}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
